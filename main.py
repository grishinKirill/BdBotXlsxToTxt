import openpyxl
import time
from pathlib import Path

xlsx_file = Path('file.xlsx')
wb_obj = openpyxl.load_workbook(xlsx_file)
ws = wb_obj.active


    

col_names = []
#print(wb_obj.active)
#print(sheet.max_raw)
#for column in sheet.iter_cols(1, sheet.max_column):
    #col_names.append(column[0].value)
   # print(column[0].value)
#current_time = time.ctime()
#print(time.strftime("%Y_%M_%d"))


current_time = time.strftime('%Y-%m-%d_%H:%M:%S.txt')
print(current_time)
with open(current_time, 'w') as f:
#with open('result.txt', 'w') as f:
    for row in ws.values:
        #line=row[0].tostring+','
        str=f'{row[0]},{row[3]};\n'
        #print(str)
        #+row[3]+';'
#        print(line)
        
        f.write(str)
   #for value in row:
     #print(value)   
f.close()
